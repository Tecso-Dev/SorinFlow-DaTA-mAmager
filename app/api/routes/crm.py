"""
SorinFlow CRM — API routes
Leads (from scraper) + Contacts + Notes + Tasks + Deals + Reminders + SMS + Dashboard
"""
from datetime import datetime, timedelta
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, and_, not_
import io
from loguru import logger

from app.database import get_db
from app.config import get_settings
from app.models.lead import Lead
from app.models.property import Property, allocate_serial_no
from app.models.crm_models import (
    Contact, Deal, Note, Task, Reminder, SmsLog, Customer, DailyPerformance,
    ActivityLog, CalendarEvent,
)
from app.schemas import LeadResponse, LeadUpdate, LeadCreate, LeadList
from app.crm.notification import notify
from app.services.sms_service import send_sms
from app.auth.dependencies import get_current_user, get_current_user_optional, require_super_admin
from app.services.dpa_service import record_activity, record_lead_status
from app.services.excel_export import xlsx_response, fa_date
from app.services.match_service import (
    similar_to_property, matches_for_customer, customer_intent, customers_for_property,
)
from app.scraper.parsers import infer_advertiser_type
from app.services.property_kind import PROPERTY_KINDS, kind_options
from app.models.user import User

router = APIRouter()

# ─────────────────────────────────────────────────────────────────────────────
# LEADS (existing — from scraper)
# ─────────────────────────────────────────────────────────────────────────────

VALID_LEAD_STATUSES = {"new", "contacted", "visit", "contract_meeting", "qualified", "closed", "rented", "rejected"}

# ── structured attributes for manual leads ──
# ints/bools/text mapped to real Property columns; the rest live in extra_attrs
_ATTR_INT_COLS  = {"area", "floor", "rooms", "year_built", "land_area", "built_area", "total_floors", "frontage"}
_ATTR_BOOL_COLS = {"has_elevator", "has_parking", "has_storage", "has_balcony"}
# text column → its own max length, so a long paste cannot overflow the column
_ATTR_TEXT_COLS = {"document_type": 100, "building_direction": 50, "corner_type": 20}
_ATTR_EXTRA_KEYS = {"units_per_floor", "cabinets", "closet", "flooring", "delivery_date",
                    "hvac", "yard", "position", "height", "mezzanine", "kitchen"}
VALID_PROPERTY_KINDS = {"apartment", "villa", "shop", "office"}


def _split_lead_attrs(attrs: Optional[dict]):
    """Split incoming attrs into Property column kwargs + extra_attrs JSON."""
    cols, extra = {}, {}
    for key, val in (attrs or {}).items():
        if val is None or val == "":
            continue
        if key in _ATTR_INT_COLS:
            try:
                cols[key] = int(val)
            except (TypeError, ValueError):
                continue
        elif key in _ATTR_BOOL_COLS:
            cols[key] = str(val).lower() in ("true", "1", "yes", "بله")
        elif key in _ATTR_TEXT_COLS:
            cols[key] = str(val).strip()[:_ATTR_TEXT_COLS[key]]
        elif key in _ATTR_EXTRA_KEYS:
            extra[key] = str(val).strip()[:200]
    return cols, extra


def _apply_lead_filters(
    query,
    *,
    status=None, city=None, category=None, search=None, notified=None,
    date_from=None, date_to=None, price_min=None, price_max=None,
    property_kind=None,
):
    """Every filter the leads list understands, in one place.

    The Excel export used to re-implement only «status», so exporting a
    filtered view handed back rows the screen was not showing. Both callers go
    through this now, and neither can drift from the other.
    """
    from datetime import datetime, timedelta

    if status:
        query = query.where(Lead.status == status)
    if city:
        query = query.where(Lead.city_name == city)
    if category:
        query = query.where(Lead.category_name.ilike(f"%{category}%"))
    if search:
        raw = search.strip()
        term = f"%{raw}%"
        # Street/neighborhood names usually live in the linked property's
        # address or description, not in the lead's own columns
        prop_filters = [
            Property.address.ilike(term),
            Property.district.ilike(term),
            Property.neighborhood.ilike(term),
            Property.description.ilike(term),
        ]
        # the leads table shows کد ملک, so a bare number should find that lead
        if raw.isdigit():
            prop_filters.append(Property.serial_no == int(raw))
        prop_match = select(Property.id).where(
            Property.id == Lead.property_id,
            or_(*prop_filters),
        ).correlate(Lead).exists()
        query = query.where(or_(
            Lead.property_title.ilike(term),
            Lead.notes.ilike(term),
            Lead.phone_number.ilike(term),
            Lead.seller_name.ilike(term),
            prop_match,
        ))
    if notified is not None:
        query = query.where(Lead.notified == notified)
    # Lead.price already holds the headline number — the sale price, or the
    # deposit on a rental — so one band works across both kinds of listing.
    if price_min is not None:
        query = query.where(Lead.price >= price_min)
    if price_max is not None:
        query = query.where(Lead.price <= price_max)
    # ── نوع ملک ──────────────────────────────────────────────────────────
    # «خرید مسکونی» on Divar holds apartments, villas and کلنگی together, so
    # the category alone cannot separate them. The kind lives on the linked
    # property: as Persian text when the scraper read it off the breadcrumb, as
    # an English slug when the lead was entered by hand, and sometimes not at
    # all — in which case the ad's own title is the only thing left to go on.
    spec = PROPERTY_KINDS.get((property_kind or "").strip())
    if spec:
        stored = [Property.property_type.ilike(v) for v in spec["values"]]
        title_any = [Property.title.ilike(f"%{w}%") for w in spec["title_any"]]
        title_no = [Property.title.ilike(f"%{w}%") for w in spec["title_none"]]
        by_title = and_(Property.property_type.is_(None), or_(*title_any))
        if title_no:
            by_title = and_(by_title, not_(or_(*title_no)))
        query = query.where(
            select(Property.id)
            .where(Property.id == Lead.property_id, or_(or_(*stored), by_title))
            .correlate(Lead).exists()
        )
    if date_from:
        try:
            query = query.where(Lead.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to)
            # include the full end day
            if len(date_to) == 10:
                dt_to = dt_to + timedelta(days=1)
            query = query.where(Lead.created_at < dt_to)
        except ValueError:
            pass

    return query


@router.get("/leads", response_model=LeadList)
async def list_leads(
    status: Optional[str] = None,
    city: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    notified: Optional[bool] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    price_min: Optional[int] = None,
    price_max: Optional[int] = None,
    property_kind: Optional[str] = None,
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    query = _apply_lead_filters(
        select(Lead).order_by(Lead.created_at.desc()),
        status=status, city=city, category=category, search=search,
        notified=notified, date_from=date_from, date_to=date_to,
        price_min=price_min, price_max=price_max, property_kind=property_kind,
    )

    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar_one()
    result = await db.execute(query.limit(limit).offset(offset))
    leads = result.scalars().all()
    return LeadList(items=await _attach_property_columns(db, leads), total=total)


@router.post("/leads", response_model=LeadResponse)
async def create_lead(
    data: LeadCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    if data.status and data.status not in VALID_LEAD_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")

    import uuid
    manual_id = f"manual-{uuid.uuid4().hex[:12]}"

    # Manually-added leads aren't scraped from Divar, so we create a minimal
    # Property row to satisfy Lead.property_id's NOT NULL FK.
    is_rent = data.listing_type == "rent"
    attr_cols, extra_attrs = _split_lead_attrs(data.attrs)
    images = [u for u in (data.images or []) if isinstance(u, str) and u.startswith("/images/")][:20]
    prop = Property(
        tag_number=manual_id,
        divar_id=manual_id,
        title=data.property_title,
        city_name=data.city_name,
        category_name=data.category_name,
        listing_type=data.listing_type,
        property_type=data.property_kind if data.property_kind in VALID_PROPERTY_KINDS else None,
        price=None if is_rent else data.price,
        total_price=None if is_rent else data.price,
        deposit=data.deposit if is_rent else None,
        rent_price=data.rent_price if is_rent else None,
        area=data.area,
        phone_number=data.phone_number,
        seller_name=data.seller_name,
        url=data.property_url or "",
        images=images,
        thumbnail_url=images[0] if images else None,
        has_images=bool(images),
        # «مهندس فلانی» or «املاک آرین» is a shop, whatever the form was told
        advertiser_type=infer_advertiser_type(data.seller_name),
        extra_attrs=extra_attrs,
        owner_phone=current_user.divar_phone if current_user else None,
        **attr_cols,
    )
    prop.serial_no = await allocate_serial_no(db)
    db.add(prop)
    await db.flush()

    lead = Lead(
        property_id=prop.id,
        phone_number=data.phone_number,
        seller_name=data.seller_name,
        city_name=data.city_name,
        category_name=data.category_name,
        listing_type=data.listing_type,
        # the lead's headline number: sale price, or the deposit for rents
        price=(data.deposit or data.rent_price) if is_rent else data.price,
        area=attr_cols.get("area", data.area),
        property_url=data.property_url or "",
        property_title=data.property_title,
        status=data.status or "new",
        notes=data.notes,
        assigned_to=data.assigned_to,
    )
    db.add(lead)

    agent = (data.assigned_to or "").strip() or (
        current_user.full_name or current_user.username if current_user else None)
    await record_activity(db, agent, "new_file")

    await db.commit()
    await db.refresh(lead)
    return lead


def _derive_price_per_meter(row) -> Optional[int]:
    """Divar often omits the per-meter price on total-price ads — derive it."""
    if row.price_per_meter:
        return row.price_per_meter
    base = row.total_price or row.price
    if base and row.area:
        return int(base / row.area)
    return None


async def _attach_property_columns(db: AsyncSession, leads) -> List[LeadResponse]:
    """Fill the property-owned columns the leads table renders.

    One extra query per page instead of a join: selecting whole Property rows
    would drag description/images through the serializer for data no column
    displays.
    """
    items = [LeadResponse.model_validate(l) for l in leads]
    prop_ids = {l.property_id for l in leads if l.property_id}
    if not prop_ids:
        return items

    rows = (await db.execute(
        select(
            Property.id, Property.serial_no,
            Property.scraped_at, Property.posted_at, Property.district,
            Property.price_per_meter, Property.total_price, Property.price, Property.area,
            Property.document_type, Property.has_parking, Property.has_elevator,
            Property.building_direction, Property.corner_type,
        ).where(Property.id.in_(prop_ids))
    )).all()
    by_id = {r.id: r for r in rows}

    for item in items:
        p = by_id.get(item.property_id)
        if not p:
            continue
        # one property has one code everywhere it appears
        item.serial_no = p.serial_no
        # «برداشت» = when we pulled the ad; fall back to the posting date
        item.scraped_at = p.scraped_at or p.posted_at
        item.price_per_meter = _derive_price_per_meter(p)
        item.document_type = p.document_type
        item.has_parking = p.has_parking
        item.has_elevator = p.has_elevator
        item.building_direction = p.building_direction
        item.corner_type = p.corner_type
        item.district = item.district or p.district
    return items


async def _lead_with_property(db: AsyncSession, lead: Lead) -> LeadResponse:
    """Attach the full linked-property snapshot so the CRM lead modal can show
    exactly what the املاک modal shows."""
    resp = LeadResponse.model_validate(lead)
    prop = (await db.execute(
        select(Property).where(Property.id == lead.property_id)
    )).scalar_one_or_none()
    if prop:
        data = prop.to_dict()
        # to_dict() omits a few columns the detail view renders
        data.update({
            "serial_no": prop.serial_no,
            "land_area": prop.land_area,
            "built_area": prop.built_area,
            "price_per_meter": prop.price_per_meter,
            "total_price": prop.total_price,
            "deposit": prop.deposit,
            "rent_price": prop.rent_price,
            "building_direction": prop.building_direction,
            "frontage": prop.frontage,
            "unit_status": prop.unit_status,
            "document_type": prop.document_type,
            "usage_type": prop.usage_type,
            "building_age": prop.building_age,
            "latitude": prop.latitude,
            "longitude": prop.longitude,
            "advertiser_type": prop.advertiser_type,
            "has_images": prop.has_images,
            "extra_attrs": prop.extra_attrs or {},
            "updated_at": prop.updated_at.isoformat() if prop.updated_at else None,
        })
        resp.property_detail = data
        resp.district = prop.district
        resp.serial_no = prop.serial_no
        # same columns the list rows show, so the modal and the table agree
        resp.scraped_at = prop.scraped_at or prop.posted_at
        resp.price_per_meter = _derive_price_per_meter(prop)
        resp.document_type = prop.document_type
        resp.has_parking = prop.has_parking
        resp.has_elevator = prop.has_elevator
        resp.building_direction = prop.building_direction
        resp.corner_type = prop.corner_type
    return resp


@router.get("/leads/export/excel")
async def export_leads_excel(
    status: Optional[str] = None,
    city: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    notified: Optional[bool] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    price_min: Optional[int] = None,
    price_max: Optional[int] = None,
    property_kind: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Excel counterpart of the leads list — same filters, same rows.

    It used to honour «status» alone, so exporting a narrowed view silently
    handed back everything else too.
    """
    query = _apply_lead_filters(
        select(Lead).order_by(Lead.created_at.desc()),
        status=status, city=city, category=category, search=search,
        notified=notified, date_from=date_from, date_to=date_to,
        price_min=price_min, price_max=price_max, property_kind=property_kind,
    )
    items = (await db.execute(query.limit(5000))).scalars().all()

    status_fa = {
        "new": "جدید", "contacted": "تماس گرفته", "visit": "بازدید از فایل",
        "contract_meeting": "نشست و تنظیم قرارداد", "qualified": "واجد شرایط",
        "closed": "بسته شده", "rented": "اجاره شده", "rejected": "رد شده",
    }
    # mirror the on-screen columns, including the ones owned by the property
    enriched = await _attach_property_columns(db, items)
    headers = ["کد ملک", "عنوان ملک", "شهر", "دسته‌بندی", "قیمت", "قیمت هر متر", "متراژ",
               "سند", "پارکینگ", "آسانسور", "جهت", "نبش",
               "شماره تماس", "فروشنده", "وضعیت CRM", "مسئول پیگیری",
               "اطلاع‌رسانی", "یادداشت", "تاریخ برداشت آگهی", "تاریخ ثبت"]
    yn = lambda v: "" if v is None else ("دارد" if v else "ندارد")
    rows = [[
        l.serial_no, l.property_title, l.city_name, l.category_name, l.price,
        l.price_per_meter, l.area,
        l.document_type, yn(l.has_parking), yn(l.has_elevator),
        l.building_direction, l.corner_type,
        l.phone_number, l.seller_name, status_fa.get(l.status, l.status),
        l.assigned_to, "بله" if l.notified else "خیر", l.notes,
        fa_date(l.scraped_at), fa_date(l.created_at),
    ] for l in enriched]
    return xlsx_response("leads.xlsx", "لیدها", headers, rows)


@router.post("/leads/bulk")
async def bulk_update_leads(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    """Bulk status change or delete for the selected leads.

    body: {"ids": [1,2,3], "action": "status"|"delete", "status": "contacted"}
    """
    ids = [int(i) for i in (data.get("ids") or []) if str(i).isdigit()][:500]
    action = data.get("action")
    if not ids:
        raise HTTPException(status_code=400, detail="هیچ لیدی انتخاب نشده است")

    leads = (await db.execute(select(Lead).where(Lead.id.in_(ids)))).scalars().all()

    if action == "status":
        new_status = data.get("status")
        if new_status not in VALID_LEAD_STATUSES:
            raise HTTPException(status_code=400, detail="Invalid status")
        changed = 0
        for lead in leads:
            if lead.status != new_status:
                lead.status = new_status
                if new_status == "rented":
                    lead.rented_at = datetime.now()
                elif lead.rented_at:
                    lead.rented_at = None
                agent = (lead.assigned_to or "").strip() or (
                    current_user.full_name or current_user.username if current_user else None)
                await record_lead_status(db, agent, new_status)
                _log_activity(db, "lead", lead.id, "status_change",
                              f"وضعیت گروهی به «{new_status}» تغییر کرد", agent)
                changed += 1
        await db.commit()
        return {"success": True, "updated": changed}

    if action == "delete":
        removed = 0
        for lead in leads:
            await db.delete(lead)
            removed += 1
        await db.commit()
        return {"success": True, "deleted": removed}

    raise HTTPException(status_code=400, detail="Unknown action")


@router.get("/leads/{lead_id}", response_model=LeadResponse)
async def get_lead(lead_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return await _lead_with_property(db, lead)


@router.patch("/leads/{lead_id}", response_model=LeadResponse)
async def update_lead(
    lead_id: int,
    data: LeadUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    if data.status and data.status not in VALID_LEAD_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status")
    if data.status is not None:
        # Only score a real transition, so re-saving the same status can't farm points
        status_changed = lead.status != data.status
        lead.status = data.status
        if status_changed and data.status == "rented":
            lead.rented_at = datetime.now()   # lease clock starts; back in a year
        elif status_changed and lead.rented_at:
            lead.rented_at = None
        if status_changed:
            agent = (lead.assigned_to or "").strip() or (
                current_user.full_name or current_user.username if current_user else None)
            await record_lead_status(db, agent, data.status)
            _log_activity(db, "lead", lead.id, "status_change",
                          f"وضعیت به «{data.status}» تغییر کرد", agent)
    if data.notes is not None:
        lead.notes = data.notes
    if data.assigned_to is not None:
        lead.assigned_to = data.assigned_to
    if data.district is not None:
        # District belongs to the linked property (street search reads it there)
        prop = (await db.execute(
            select(Property).where(Property.id == lead.property_id)
        )).scalar_one_or_none()
        if prop:
            prop.district = data.district.strip() or None
    await db.commit()
    await db.refresh(lead)
    return await _lead_with_property(db, lead)


@router.post("/leads/{lead_id}/notify")
async def notify_lead(lead_id: int, db: AsyncSession = Depends(get_db)):
    lead_result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = lead_result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    prop_result = await db.execute(select(Property).where(Property.id == lead.property_id))
    prop = prop_result.scalar_one_or_none()
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    channel = await notify(prop, lead)
    lead.notified = channel != "none"
    lead.notified_at = datetime.now() if lead.notified else lead.notified_at
    lead.notification_channel = channel
    await db.commit()
    return {"success": lead.notified, "channel": channel}


@router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: int, db: AsyncSession = Depends(get_db)):
    """Delete the lead AND wipe the linked property everywhere:
    the property row itself, any sibling leads on it, its property-notes
    and its downloaded images. Deals keep their business record (their
    property link is nulled by the FK)."""
    result = await db.execute(select(Lead).where(Lead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    property_id = lead.property_id
    prop = (await db.execute(
        select(Property).where(Property.id == property_id)
    )).scalar_one_or_none()

    # all leads pointing at this property (FK is NOT NULL, so they must go
    # before the property row can)
    siblings = (await db.execute(
        select(Lead).where(Lead.property_id == property_id)
    )).scalars().all()
    for sib in siblings:
        await db.delete(sib)

    # property-specific notes
    notes = (await db.execute(
        select(Note).where(Note.property_id == property_id)
    )).scalars().all()
    for note in notes:
        await db.delete(note)

    divar_id = prop.divar_id if prop else None
    if prop:
        await db.delete(prop)
    await db.commit()

    # downloaded images on disk (best-effort)
    if divar_id:
        try:
            import shutil
            from app.config import get_settings as _gs
            img_dir = __import__("pathlib").Path(_gs().images_path) / divar_id
            if img_dir.is_dir():
                shutil.rmtree(img_dir, ignore_errors=True)
        except Exception:
            pass

    return {"success": True, "deleted_property": prop is not None,
            "deleted_leads": len(siblings), "deleted_notes": len(notes)}


@router.post("/upload-image")
async def upload_lead_image(file: UploadFile = File(...)):
    """Store a lead photo (converted to JPEG) under data/images/manual."""
    import uuid as _uuid
    import io as _io
    from pathlib import Path as _Path
    from PIL import Image as _Image

    if not (file.content_type or "").startswith("image/"):
        raise HTTPException(status_code=400, detail="فقط فایل تصویری مجاز است")
    raw = await file.read()
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="حجم تصویر حداکثر ۸ مگابایت")

    # Always convert (webp/png/heic-ish) to JPEG for a universal format
    try:
        im = _Image.open(_io.BytesIO(raw)).convert("RGB")
    except Exception:
        raise HTTPException(status_code=400, detail="فایل تصویری معتبر نیست")

    dest_dir = _Path(get_settings().images_path) / "manual"
    dest_dir.mkdir(parents=True, exist_ok=True)
    name = f"{_uuid.uuid4().hex}.jpg"
    im.save(dest_dir / name, format="JPEG", quality=88)
    return {"url": f"/images/manual/{name}"}


@router.get("/customers/export/excel")
async def export_customers_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Excel export of the customer intake list."""
    items = (await db.execute(
        select(Customer).order_by(Customer.created_at.desc()).limit(5000)
    )).scalars().all()

    temp_fa = {"hot": "داغ", "warm": "گرم", "cold": "سرد"}
    source_fa = {"in_person": "حضوری", "divar": "دیوار", "referral": "معرف"}
    type_fa = {"apartment": "آپارتمان", "house": "ویلایی / خانه", "land": "زمین",
               "shop": "مغازه", "office": "دفتر کار"}
    deal_fa = {"buy": "خرید", "rent": "رهن و اجاره"}
    headers = ["#", "نام و نام خانوادگی", "موبایل ۱", "موبایل ۲", "حرارت", "منبع",
               "مشاور", "نوع معامله", "نوع ملک", "شهر", "سقف بودجه", "نحوه پرداخت",
               "متراژ/خواب", "منطقه درخواستی",
               "خط قرمزها", "تعداد بازدید", "پیگیری بعدی", "یادداشت", "تاریخ ثبت"]
    rows = []
    for c in items:
        followups = c.followups or []
        next_fu = f"{followups[0].get('date','')} {followups[0].get('time','')}".strip() if followups else ""
        rows.append([
            c.id, c.full_name, c.mobile1, c.mobile2,
            temp_fa.get(c.temperature, c.temperature), source_fa.get(c.source, c.source),
            c.consultant_name,
            deal_fa.get(c.deal_type or "buy", ""), type_fa.get(c.desired_type, ""), c.desired_city,
            c.budget_max, c.payment_methods, c.desired_specs,
            c.desired_district, c.red_lines, len(c.showings or []), next_fu,
            c.notes, fa_date(c.created_at),
        ])
    return xlsx_response("customers.xlsx", "مشتریان", headers, rows)


@router.get("/dpa/export/excel")
async def export_dpa_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Excel export of daily performance records with the score breakdown."""
    items = (await db.execute(
        select(DailyPerformance).order_by(DailyPerformance.created_at.desc()).limit(5000)
    )).scalars().all()

    headers = ["#", "تاریخ", "مشاور", "نقش", "فایل جدید", "بازدید", "آفر", "قرارداد",
               "امتیاز پایه", "امتیاز فعالیت", "بونوس", "جریمه", "امتیاز کل", "هدف",
               "تحلیل موانع", "بازخورد منتور"]
    rows = []
    for d in items:
        s = d.scores()
        rows.append([
            d.id, d.date_jalali, d.agent_name, d.role, d.new_files, d.showings_count,
            d.offers_count, d.closed_count, s["base_score"], s["activity_score"],
            s["bonus_score"], s["penalty_score"], s["total_score"], d.target_points,
            d.rca, d.mentor_feedback,
        ])
    return xlsx_response("dpa.xlsx", "ارزیابی روزانه", headers, rows)


def _log_activity(db: AsyncSession, entity_type: str, entity_id: int,
                  action: str, detail: str, actor: Optional[str] = None) -> None:
    """Append a timeline entry. Fire-and-forget: never breaks the caller."""
    try:
        db.add(ActivityLog(entity_type=entity_type, entity_id=entity_id,
                           action=action, detail=detail[:500], actor=actor))
    except Exception as e:
        logger.warning(f"[activity] failed to log {action}: {e}")


@router.get("/activity/{entity_type}/{entity_id}")
async def get_activity(
    entity_type: str,
    entity_id: int,
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Timeline for a lead / customer / deal."""
    if entity_type not in ("lead", "customer", "deal"):
        raise HTTPException(status_code=400, detail="Unknown entity type")
    rows = (await db.execute(
        select(ActivityLog)
        .where(ActivityLog.entity_type == entity_type, ActivityLog.entity_id == entity_id)
        .order_by(ActivityLog.created_at.desc()).limit(limit)
    )).scalars().all()
    return {"items": [r.to_dict() for r in rows], "total": len(rows)}


@router.post("/leads/{lead_id}/convert-to-deal")
async def convert_lead_to_deal(
    lead_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    """Create a deal pre-filled from this lead (title, property, amount, seller)."""
    lead = (await db.execute(select(Lead).where(Lead.id == lead_id))).scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    # reuse an existing contact with the same phone, otherwise create one
    seller_id = None
    if lead.phone_number:
        existing = (await db.execute(
            select(Contact).where(Contact.phone == lead.phone_number)
        )).scalars().first()
        if existing:
            seller_id = existing.id
        else:
            contact = Contact(
                name=lead.seller_name or f"فروشنده {lead.property_title or ''}".strip(),
                phone=lead.phone_number, contact_type="owner",
                city=lead.city_name, notes=f"از لید #{lead.id} ساخته شد",
            )
            db.add(contact)
            await db.flush()
            seller_id = contact.id

    deal = Deal(
        title=lead.property_title or f"معامله لید #{lead.id}",
        deal_type="rent" if lead.listing_type == "rent" else "buy",
        status="new",
        property_id=lead.property_id,
        seller_contact_id=seller_id,
        amount=lead.price,
        notes=f"ساخته‌شده از لید #{lead.id}" + (f"\n{lead.notes}" if lead.notes else ""),
    )
    db.add(deal)
    await db.flush()

    actor = (lead.assigned_to or "").strip() or (
        current_user.full_name or current_user.username if current_user else None)
    _log_activity(db, "lead", lead.id, "converted",
                  f"تبدیل به معامله #{deal.id}", actor)
    _log_activity(db, "deal", deal.id, "created",
                  f"ساخته‌شده از لید #{lead.id}", actor)
    await db.commit()
    await db.refresh(deal)
    return {"success": True, "deal": deal.to_dict()}


# ─────────────────────────────────────────────────────────────────────────────
# MATCHING — «تطابق‌سازی»
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/match/property/{property_id}")
async def match_similar_properties(
    property_id: int,
    limit: int = Query(12, ge=1, le=50),
    use_llm: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Listings similar to this one — «مشتری این ملک را پسندید، مشابهش را نشان بده»."""
    prop = (await db.execute(select(Property).where(Property.id == property_id))).scalar_one_or_none()
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    items = await similar_to_property(db, prop, limit=limit, use_llm=use_llm)
    return {"items": items, "total": len(items),
            "source": {"id": prop.id, "title": prop.title, "serial_no": prop.serial_no}}


@router.get("/match/lead/{lead_id}")
async def match_similar_for_lead(
    lead_id: int,
    limit: int = Query(12, ge=1, le=50),
    use_llm: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Same as above but addressed by lead id (the CRM's natural handle)."""
    lead = (await db.execute(select(Lead).where(Lead.id == lead_id))).scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    prop = (await db.execute(select(Property).where(Property.id == lead.property_id))).scalar_one_or_none()
    if not prop:
        raise HTTPException(status_code=404, detail="Linked property not found")
    items = await similar_to_property(db, prop, limit=limit, use_llm=use_llm)
    return {"items": items, "total": len(items),
            "source": {"id": prop.id, "title": prop.title, "serial_no": prop.serial_no}}


@router.get("/match/customer/{customer_id}")
async def match_properties_for_customer(
    customer_id: int,
    limit: int = Query(12, ge=1, le=50),
    use_llm: bool = True,
    city: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Listings that fit this customer's budget / district / specs (BANT).

    `city` narrows the search — the intake form has no city field, and street
    names repeat across cities, so without it a matching district elsewhere
    can surface.
    """
    customer = (await db.execute(select(Customer).where(Customer.id == customer_id))).scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    items = await matches_for_customer(db, customer, limit=limit, use_llm=use_llm, city=city)
    return {"items": items, "total": len(items),
            "source": {"id": customer.id, "name": customer.full_name},
            "intent": customer_intent(customer)}


# ─────────────────────────────────────────────────────────────────────────────
# CONTACTS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/contacts")
async def list_contacts(
    search: Optional[str] = None,
    contact_type: Optional[str] = None,
    category: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    query = select(Contact).order_by(Contact.created_at.desc())
    if search:
        query = query.where(or_(
            Contact.name.ilike(f"%{search}%"),
            Contact.phone.ilike(f"%{search}%"),
            Contact.phone2.ilike(f"%{search}%"),
        ))
    if contact_type:
        query = query.where(Contact.contact_type == contact_type)
    if category:
        query = query.where(Contact.category == category)
    count = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    items = (await db.execute(query.limit(limit).offset(offset))).scalars().all()
    return {"items": [c.to_dict() for c in items], "total": count}


def _normalize_tags(tags) -> str | None:
    if tags is None:
        return None
    if isinstance(tags, list):
        return ", ".join(str(t).strip() for t in tags if t)
    return str(tags).strip() or None


@router.post("/contacts")
async def create_contact(data: dict, db: AsyncSession = Depends(get_db)):
    contact = Contact(
        name=data.get("name", ""),
        phone=data.get("phone"),
        phone2=data.get("phone2"),
        email=data.get("email"),
        contact_type=data.get("contact_type", "owner"),
        category=data.get("category", "normal"),
        city=data.get("city"),
        address=data.get("address"),
        notes=data.get("notes"),
        tags=_normalize_tags(data.get("tags")),
    )
    db.add(contact)
    await db.commit()
    await db.refresh(contact)
    return contact.to_dict()


@router.get("/contacts/export/excel")
async def export_contacts_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    items = (await db.execute(select(Contact).order_by(Contact.name))).scalars().all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "مخاطبین"
    headers = ["نام", "تلفن", "تلفن ۲", "ایمیل", "دسته", "اولویت", "شهر", "آدرس", "تگ‌ها", "تاریخ ثبت"]
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    contact_type_fa = {
        "owner": "مالکین", "landlord": "موجرین", "tenant": "مستاجرین",
        "seeker": "خواهان", "builder": "سازندگان", "agency": "املاک",
        "buyer": "خواهان", "seller": "مالکین", "consultant": "املاک", "other": "سایر",
    }
    for row, c in enumerate(items, 2):
        ws.append([
            c.name, c.phone, c.phone2, c.email,
            contact_type_fa.get(c.contact_type, c.contact_type), c.category, c.city, c.address,
            ", ".join(c.tags) if isinstance(c.tags, list) else (c.tags or ""),
            c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contacts.xlsx"},
    )


@router.get("/contacts/export/json")
async def export_contacts_json(
    db: AsyncSession = Depends(get_db),
    current_user: User = require_super_admin,
):
    items = (await db.execute(select(Contact).order_by(Contact.name))).scalars().all()
    import json
    data = json.dumps([c.to_dict() for c in items], ensure_ascii=False, default=str, indent=2)
    return StreamingResponse(
        io.BytesIO(data.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=contacts.json"},
    )


@router.get("/contacts/{contact_id}")
async def get_contact(contact_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Contact).where(Contact.id == contact_id))
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    return contact.to_dict()


@router.put("/contacts/{contact_id}")
async def update_contact(contact_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Contact).where(Contact.id == contact_id))
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    for field in ("name", "phone", "phone2", "email", "contact_type", "category", "city", "address", "notes"):
        if field in data:
            setattr(contact, field, data[field])
    if "tags" in data:
        contact.tags = _normalize_tags(data["tags"])
    contact.updated_at = datetime.now()
    await db.commit()
    await db.refresh(contact)
    return contact.to_dict()


@router.delete("/contacts/{contact_id}")
async def delete_contact(contact_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Contact).where(Contact.id == contact_id))
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    await db.delete(contact)
    await db.commit()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOMERS (فرم پروفایل مشتری: BANT + بازدیدها + پیگیری)
# ─────────────────────────────────────────────────────────────────────────────

VALID_CUSTOMER_SOURCES = {"in_person", "divar", "referral"}
VALID_CUSTOMER_TEMPS = {"hot", "warm", "cold"}
VALID_CUSTOMER_TYPES = {"apartment", "house", "land", "shop", "office"}
VALID_DEAL_TYPES = {"buy", "rent"}
CUSTOMER_TEXT_FIELDS = (
    "full_name", "mobile1", "mobile2", "consultant_name",
    "payment_methods", "desired_specs", "desired_district", "desired_city",
    "red_lines", "notes",
)


def _clean_customer_rows(rows, keys) -> list:
    """Keep only known keys per row; drop rows that are entirely empty."""
    out = []
    for row in (rows or []):
        if not isinstance(row, dict):
            continue
        cleaned = {k: (str(row.get(k) or "").strip() or None) for k in keys}
        if any(cleaned.values()):
            out.append(cleaned)
    return out


def _apply_customer_payload(customer: Customer, data: dict) -> None:
    for field in CUSTOMER_TEXT_FIELDS:
        if field in data:
            setattr(customer, field, (str(data[field]).strip() or None) if data[field] is not None else None)
    if "source" in data and data["source"] in VALID_CUSTOMER_SOURCES:
        customer.source = data["source"]
    if "temperature" in data and data["temperature"] in VALID_CUSTOMER_TEMPS:
        customer.temperature = data["temperature"]
    if "desired_type" in data:
        # "" clears it and puts the matcher back on text inference
        customer.desired_type = data["desired_type"] if data["desired_type"] in VALID_CUSTOMER_TYPES else None
    if "deal_type" in data and data["deal_type"] in VALID_DEAL_TYPES:
        customer.deal_type = data["deal_type"]
    if "budget_max" in data:
        try:
            customer.budget_max = int(data["budget_max"]) if data["budget_max"] else None
        except (TypeError, ValueError):
            pass
    if "showings" in data:
        customer.showings = _clean_customer_rows(
            data["showings"], ("file_code", "description", "feedback", "next_step"))
    if "followups" in data:
        customer.followups = _clean_customer_rows(
            data["followups"], ("date", "time", "action"))


@router.get("/customers")
async def list_customers(
    search: Optional[str] = None,
    temperature: Optional[str] = None,
    source: Optional[str] = None,
    sort: str = "newest",
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    order = {
        "newest": Customer.created_at.desc(),
        "oldest": Customer.created_at.asc(),
        "name": Customer.full_name.asc(),
    }.get(sort, Customer.created_at.desc())
    query = select(Customer).order_by(order)
    if search:
        term = f"%{search.strip()}%"
        query = query.where(or_(
            Customer.full_name.ilike(term),
            Customer.mobile1.ilike(term),
            Customer.mobile2.ilike(term),
            Customer.desired_district.ilike(term),
            Customer.consultant_name.ilike(term),
        ))
    if temperature:
        query = query.where(Customer.temperature == temperature)
    if source:
        query = query.where(Customer.source == source)
    count = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    items = (await db.execute(query.limit(limit).offset(offset))).scalars().all()
    return {"items": [c.to_dict() for c in items], "total": count}


@router.post("/customers")
async def create_customer(data: dict, db: AsyncSession = Depends(get_db)):
    if not str(data.get("full_name") or "").strip():
        raise HTTPException(status_code=400, detail="full_name is required")
    customer = Customer(full_name=str(data["full_name"]).strip())
    _apply_customer_payload(customer, data)
    db.add(customer)
    await db.commit()
    await db.refresh(customer)
    return customer.to_dict()


@router.get("/customers/{customer_id}")
async def get_customer(customer_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Customer).where(Customer.id == customer_id))
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer.to_dict()


@router.put("/customers/{customer_id}")
async def update_customer(customer_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Customer).where(Customer.id == customer_id))
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    _apply_customer_payload(customer, data)
    customer.updated_at = datetime.now()
    await db.commit()
    await db.refresh(customer)
    return customer.to_dict()


@router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Customer).where(Customer.id == customer_id))
    customer = result.scalar_one_or_none()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    await db.delete(customer)
    await db.commit()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# DAILY PERFORMANCE (فرم ارزیابی عملکرد روزانه — DPA)
# ─────────────────────────────────────────────────────────────────────────────

_DPA_INT_FIELDS = (
    "target_points", "new_files", "showings_count", "offers_count", "closed_count",
    "bonus_exclusive", "bonus_offer", "bonus_close",
    "pen_crm_delay", "pen_cancel", "pen_hot_lead",
)
_DPA_TEXT_FIELDS = ("agent_name", "date_jalali", "mentor_feedback", "rca")
_DPA_TASK_KEYS = {key for key, _ in DailyPerformance.BASE_TASKS}


def _apply_dpa_payload(dpa: DailyPerformance, data: dict) -> None:
    for field in _DPA_TEXT_FIELDS:
        if field in data:
            setattr(dpa, field, (str(data[field]).strip() or None) if data[field] is not None else None)
    if data.get("role") in ("hunter", "closer"):
        dpa.role = data["role"]
    for field in _DPA_INT_FIELDS:
        if field in data:
            try:
                setattr(dpa, field, max(int(data[field] or 0), 0))
            except (TypeError, ValueError):
                pass
    if isinstance(data.get("base_tasks"), dict):
        dpa.base_tasks = {k: bool(v) for k, v in data["base_tasks"].items() if k in _DPA_TASK_KEYS}
    # manual activity units (work done outside the panel); auto counts are
    # owned by the system and never overwritten from the client
    if isinstance(data.get("activities"), dict):
        cleaned = {}
        for k, v in data["activities"].items():
            if k in DailyPerformance.ACTIVITY_POINTS:
                try:
                    cleaned[k] = max(int(v or 0), 0)
                except (TypeError, ValueError):
                    continue
        dpa.activities = cleaned


@router.get("/dpa")
async def list_dpa(
    search: Optional[str] = None,
    date_jalali: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    query = select(DailyPerformance).order_by(DailyPerformance.created_at.desc())
    if search:
        query = query.where(DailyPerformance.agent_name.ilike(f"%{search.strip()}%"))
    if date_jalali:
        query = query.where(DailyPerformance.date_jalali == date_jalali.strip())
    count = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    items = (await db.execute(query.limit(limit).offset(offset))).scalars().all()
    return {"items": [d.to_dict() for d in items], "total": count}


@router.post("/dpa")
async def create_dpa(data: dict, db: AsyncSession = Depends(get_db)):
    if not str(data.get("agent_name") or "").strip():
        raise HTTPException(status_code=400, detail="agent_name is required")
    dpa = DailyPerformance(agent_name=str(data["agent_name"]).strip())
    _apply_dpa_payload(dpa, data)
    db.add(dpa)
    await db.commit()
    await db.refresh(dpa)
    return dpa.to_dict()


@router.get("/dpa/{dpa_id}")
async def get_dpa(dpa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DailyPerformance).where(DailyPerformance.id == dpa_id))
    dpa = result.scalar_one_or_none()
    if not dpa:
        raise HTTPException(status_code=404, detail="DPA record not found")
    return dpa.to_dict()


@router.put("/dpa/{dpa_id}")
async def update_dpa(dpa_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DailyPerformance).where(DailyPerformance.id == dpa_id))
    dpa = result.scalar_one_or_none()
    if not dpa:
        raise HTTPException(status_code=404, detail="DPA record not found")
    _apply_dpa_payload(dpa, data)
    dpa.updated_at = datetime.now()
    await db.commit()
    await db.refresh(dpa)
    return dpa.to_dict()


@router.delete("/dpa/{dpa_id}")
async def delete_dpa(dpa_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(DailyPerformance).where(DailyPerformance.id == dpa_id))
    dpa = result.scalar_one_or_none()
    if not dpa:
        raise HTTPException(status_code=404, detail="DPA record not found")
    await db.delete(dpa)
    await db.commit()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# NOTES
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/notes")
async def list_notes(
    contact_id: Optional[int] = None,
    deal_id: Optional[int] = None,
    property_id: Optional[int] = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    query = select(Note).order_by(Note.created_at.desc())
    if contact_id:
        query = query.where(Note.contact_id == contact_id)
    if deal_id:
        query = query.where(Note.deal_id == deal_id)
    if property_id:
        query = query.where(Note.property_id == property_id)
    count = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    items = (await db.execute(query.limit(limit).offset(offset))).scalars().all()
    return {"items": [n.to_dict() for n in items], "total": count}


@router.post("/notes")
async def create_note(data: dict, db: AsyncSession = Depends(get_db)):
    note = Note(
        content=data.get("content", ""),
        contact_id=data.get("contact_id"),
        property_id=data.get("property_id"),
        deal_id=data.get("deal_id"),
        created_by=data.get("created_by"),
    )
    db.add(note)
    await db.commit()
    await db.refresh(note)
    return note.to_dict()


@router.put("/notes/{note_id}")
async def update_note(note_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Note).where(Note.id == note_id))
    note = result.scalar_one_or_none()
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    if "content" in data:
        note.content = data["content"]
    note.updated_at = datetime.now()
    await db.commit()
    await db.refresh(note)
    return note.to_dict()


@router.delete("/notes/{note_id}")
async def delete_note(note_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Note).where(Note.id == note_id))
    note = result.scalar_one_or_none()
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    await db.delete(note)
    await db.commit()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# TASKS
# ─────────────────────────────────────────────────────────────────────────────

# ── وظایف: who may see which task ───────────────────────────────────────
def _task_actor(user) -> Optional[str]:
    """The name tasks are assigned under — the same string the task form puts
    in assigned_to."""
    return getattr(user, "full_name", None) or getattr(user, "username", None)


def _tasks_visible_to(query, user):
    """A super_admin sees the whole board; everyone else sees only their own.

    Tasks with no assignee stay visible to all, because they predate this rule
    and hiding them would orphan them — new tasks are stamped with their
    creator on the way in, so the unassigned set only ever shrinks.
    """
    if getattr(user, "role", None) == "super_admin":
        return query
    actor = _task_actor(user)
    if not actor:
        return query.where(Task.assigned_to.is_(None))
    return query.where(or_(Task.assigned_to.is_(None), Task.assigned_to == actor))


@router.get("/tasks")
async def list_tasks(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    contact_id: Optional[int] = None,
    deal_id: Optional[int] = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = _tasks_visible_to(
        select(Task).order_by(Task.due_date.asc().nullslast(), Task.created_at.desc()),
        current_user)
    if status:
        query = query.where(Task.status == status)
    if priority:
        query = query.where(Task.priority == priority)
    if contact_id:
        query = query.where(Task.contact_id == contact_id)
    if deal_id:
        query = query.where(Task.deal_id == deal_id)
    count = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    items = (await db.execute(query.limit(limit).offset(offset))).scalars().all()
    return {"items": [t.to_dict() for t in items], "total": count}


@router.post("/tasks")
async def create_task(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    due = None
    if data.get("due_date"):
        try:
            due = _parse_datetime(data["due_date"])
        except Exception:
            pass
    task = Task(
        title=data.get("title", ""),
        description=data.get("description"),
        due_date=due,
        priority=data.get("priority", "medium"),
        status=data.get("status", "todo"),
        contact_id=data.get("contact_id"),
        deal_id=data.get("deal_id"),
        # falls back to the creator, so a task always has someone it belongs to
        assigned_to=(data.get("assigned_to") or "").strip() or _task_actor(current_user),
    )
    db.add(task)
    await db.commit()
    await db.refresh(task)
    return task.to_dict()


async def _own_task_or_404(task_id: int, db: AsyncSession, user) -> Task:
    """One task, but only if this user is allowed to see it. Returns 404 rather
    than 403 so an id nobody may touch is indistinguishable from one that does
    not exist."""
    task = (await db.execute(_tasks_visible_to(
        select(Task).where(Task.id == task_id), user))).scalar_one_or_none()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@router.get("/tasks/{task_id}")
async def get_task(
    task_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = await _own_task_or_404(task_id, db, current_user)
    return task.to_dict()


@router.put("/tasks/{task_id}")
async def update_task(
    task_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = await _own_task_or_404(task_id, db, current_user)
    for field in ("title", "description", "priority", "status", "contact_id", "deal_id", "assigned_to"):
        if field in data:
            setattr(task, field, data[field])
    if "due_date" in data and data["due_date"]:
        try:
            task.due_date = _parse_datetime(data["due_date"])
        except Exception:
            pass
    task.updated_at = datetime.now()
    await db.commit()
    await db.refresh(task)
    return task.to_dict()


@router.patch("/tasks/{task_id}/status")
async def update_task_status(
    task_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = await _own_task_or_404(task_id, db, current_user)
    task.status = data.get("status", task.status)
    task.updated_at = datetime.now()
    await db.commit()
    return {"id": task.id, "status": task.status}


@router.delete("/tasks/{task_id}")
async def delete_task(
    task_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = await _own_task_or_404(task_id, db, current_user)
    await db.delete(task)
    await db.commit()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# DEALS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/deals")
async def list_deals(
    status: Optional[str] = None,
    deal_type: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    query = select(Deal).order_by(Deal.created_at.desc())
    if status:
        query = query.where(Deal.status == status)
    if deal_type:
        query = query.where(Deal.deal_type == deal_type)
    count = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    items = (await db.execute(query.limit(limit).offset(offset))).scalars().all()
    result = []
    for d in items:
        row = d.to_dict()
        # Fetch linked contact names
        if d.buyer_contact_id and not row.get("buyer_name"):
            r = await db.execute(select(Contact.name).where(Contact.id == d.buyer_contact_id))
            row["buyer_name"] = r.scalar_one_or_none()
        if d.seller_contact_id and not row.get("seller_name"):
            r = await db.execute(select(Contact.name).where(Contact.id == d.seller_contact_id))
            row["seller_name"] = r.scalar_one_or_none()
        result.append(row)
    return {"items": result, "total": count}


@router.post("/deals")
async def create_deal(data: dict, db: AsyncSession = Depends(get_db)):
    contract_date = close_date = None
    if data.get("contract_date"):
        try:
            contract_date = _parse_datetime(data["contract_date"])
        except Exception:
            pass
    if data.get("close_date"):
        try:
            close_date = _parse_datetime(data["close_date"])
        except Exception:
            pass
    deal = Deal(
        title=data.get("title", ""),
        deal_type=data.get("deal_type", "buy"),
        status=data.get("status", "new"),
        property_id=data.get("property_id"),
        buyer_contact_id=data.get("buyer_contact_id"),
        seller_contact_id=data.get("seller_contact_id"),
        amount=data.get("amount"),
        commission=data.get("commission"),
        commission_paid=data.get("commission_paid", False),
        notes=data.get("notes"),
        contract_date=contract_date,
        close_date=close_date,
    )
    db.add(deal)
    await db.commit()
    await db.refresh(deal)
    return deal.to_dict()


@router.get("/deals/export/excel")
async def export_deals_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    items = (await db.execute(select(Deal).order_by(Deal.created_at.desc()))).scalars().all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "معاملات"
    headers = ["عنوان", "نوع", "وضعیت", "مبلغ", "کمیسیون", "کمیسیون پرداخت شده", "تاریخ قرارداد"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
    for row, d in enumerate(items, 2):
        ws.append([
            d.title, d.deal_type, d.status, d.amount, d.commission,
            "بله" if d.commission_paid else "خیر",
            d.contract_date.strftime("%Y-%m-%d") if d.contract_date else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deals.xlsx"},
    )


@router.get("/deals/export/json")
async def export_deals_json(
    db: AsyncSession = Depends(get_db),
    current_user: User = require_super_admin,
):
    items = (await db.execute(select(Deal).order_by(Deal.created_at.desc()))).scalars().all()
    import json
    data = json.dumps([d.to_dict() for d in items], ensure_ascii=False, default=str, indent=2)
    return StreamingResponse(
        io.BytesIO(data.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=deals.json"},
    )


@router.get("/deals/{deal_id}")
async def get_deal(deal_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    return deal.to_dict()


@router.put("/deals/{deal_id}")
async def update_deal(deal_id: int, data: dict, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    for field in ("title", "deal_type", "status", "property_id", "buyer_contact_id",
                  "seller_contact_id", "amount", "commission", "commission_paid", "notes"):
        if field in data:
            setattr(deal, field, data[field])
    for date_field in ("contract_date", "close_date"):
        if data.get(date_field):
            try:
                setattr(deal, date_field, _parse_datetime(data[date_field]))
            except Exception:
                pass
    deal.updated_at = datetime.now()
    await db.commit()
    await db.refresh(deal)
    return deal.to_dict()


@router.delete("/deals/{deal_id}")
async def delete_deal(deal_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    await db.delete(deal)
    await db.commit()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# REMINDERS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/reminders")
async def list_reminders(
    is_sent: Optional[bool] = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    query = select(Reminder).order_by(Reminder.remind_at.asc())
    if is_sent is not None:
        query = query.where(Reminder.is_sent == is_sent)
    count = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    items = (await db.execute(query.limit(limit).offset(offset))).scalars().all()
    return {"items": [r.to_dict() for r in items], "total": count}


@router.get("/reminders/due")
async def get_due_reminders(db: AsyncSession = Depends(get_db)):
    """Reminders due in the next 24 hours that haven't been sent."""
    now = datetime.now()
    soon = now + timedelta(hours=24)
    result = await db.execute(
        select(Reminder).where(
            Reminder.remind_at >= now,
            Reminder.remind_at <= soon,
            Reminder.is_sent == False,
        ).order_by(Reminder.remind_at.asc())
    )
    items = result.scalars().all()
    return {"items": [r.to_dict() for r in items]}


def _parse_datetime(value: str) -> datetime:
    """Parse ISO datetime string; handles Z suffix (Python < 3.11 compat)."""
    return datetime.fromisoformat(value.replace("Z", "+00:00"))


@router.post("/reminders")
async def create_reminder(data: dict, db: AsyncSession = Depends(get_db)):
    remind_at = None
    if data.get("remind_at"):
        try:
            remind_at = _parse_datetime(data["remind_at"])
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid remind_at format")
    if not remind_at:
        raise HTTPException(status_code=400, detail="remind_at is required")
    reminder = Reminder(
        title=data.get("title", ""),
        remind_at=remind_at,
        repeat=data.get("repeat", "none"),
        channel=data.get("channel", "in_app"),
        sms_to=data.get("sms_to"),
        contact_id=data.get("contact_id"),
        deal_id=data.get("deal_id"),
        task_id=data.get("task_id"),
    )
    db.add(reminder)
    await db.commit()
    await db.refresh(reminder)
    return reminder.to_dict()


@router.delete("/reminders/{reminder_id}")
async def delete_reminder(reminder_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Reminder).where(Reminder.id == reminder_id))
    reminder = result.scalar_one_or_none()
    if not reminder:
        raise HTTPException(status_code=404, detail="Reminder not found")
    await db.delete(reminder)
    await db.commit()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# SMS
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/sms/send")
async def send_sms_route(data: dict, db: AsyncSession = Depends(get_db)):
    to_number = data.get("to_number", "").strip()
    message = data.get("message", "").strip()
    provider = data.get("provider", "kavenegar")
    contact_id = data.get("contact_id")

    if not to_number or not message:
        raise HTTPException(status_code=400, detail="to_number and message are required")

    result = await send_sms(to_number, message, provider)

    log = SmsLog(
        to_number=to_number,
        message=message,
        status="sent" if result["success"] else "failed",
        provider=provider,
        response=result.get("response", ""),
        contact_id=contact_id,
    )
    db.add(log)
    await db.commit()
    await db.refresh(log)
    return {**result, "log_id": log.id}


@router.get("/sms/logs")
async def list_sms_logs(
    contact_id: Optional[int] = None,
    status: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
):
    query = select(SmsLog).order_by(SmsLog.sent_at.desc())
    if contact_id:
        query = query.where(SmsLog.contact_id == contact_id)
    if status:
        query = query.where(SmsLog.status == status)
    count = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar_one()
    items = (await db.execute(query.limit(limit).offset(offset))).scalars().all()
    return {"items": [s.to_dict() for s in items], "total": count}


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD / STATS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/stats")
async def crm_stats(
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    # Lead stats are shared across all roles (no per-user isolation)
    def _lead_count(*extra_where):
        q = select(func.count(Lead.id))
        for w in extra_where:
            q = q.where(w)
        return q

    def _lead_group(group_col):
        return select(group_col, func.count(Lead.id)).group_by(group_col)

    # Leads
    total_leads = (await db.execute(_lead_count())).scalar_one()
    leads_by_status = {
        row[0]: row[1]
        for row in (await db.execute(_lead_group(Lead.status))).all()
    }
    notified = (await db.execute(_lead_count(Lead.notified == True))).scalar_one()

    # Contacts
    total_contacts = (await db.execute(select(func.count(Contact.id)))).scalar_one()
    contacts_by_type = {
        row[0]: row[1]
        for row in (await db.execute(select(Contact.contact_type, func.count(Contact.id)).group_by(Contact.contact_type))).all()
    }

    # Tasks
    total_tasks = (await db.execute(select(func.count(Task.id)))).scalar_one()
    tasks_todo = (await db.execute(select(func.count(Task.id)).where(Task.status == "todo"))).scalar_one()
    tasks_done = (await db.execute(select(func.count(Task.id)).where(Task.status == "done"))).scalar_one()
    now = datetime.now()
    tasks_overdue = (await db.execute(
        select(func.count(Task.id)).where(Task.due_date < now, Task.status != "done")
    )).scalar_one()

    # Deals
    total_deals = (await db.execute(select(func.count(Deal.id)))).scalar_one()
    deals_by_status = {
        row[0]: row[1]
        for row in (await db.execute(select(Deal.status, func.count(Deal.id)).group_by(Deal.status))).all()
    }
    closed_amount = (await db.execute(
        select(func.sum(Deal.amount)).where(Deal.status == "closed")
    )).scalar_one() or 0

    # Reminders due today
    today_end = now.replace(hour=23, minute=59, second=59)
    reminders_due_today = (await db.execute(
        select(func.count(Reminder.id)).where(
            Reminder.remind_at <= today_end, Reminder.is_sent == False
        )
    )).scalar_one()

    # SMS
    total_sms = (await db.execute(select(func.count(SmsLog.id)))).scalar_one()

    return {
        "leads": {
            "total": total_leads,
            "by_status": leads_by_status,
            "notified": notified,
            "pending_notification": total_leads - notified,
        },
        "contacts": {
            "total": total_contacts,
            "by_type": contacts_by_type,
        },
        "tasks": {
            "total": total_tasks,
            "todo": tasks_todo,
            "done": tasks_done,
            "overdue": tasks_overdue,
        },
        "deals": {
            "total": total_deals,
            "by_status": deals_by_status,
            "closed_amount": closed_amount,
        },
        "reminders_due_today": reminders_due_today,
        "total_sms": total_sms,
    }


# ============== Calendar — تقویم ==============
#
# The grid draws three sources at once so «برنامهٔ امروز» lives in one place:
#   • CalendarEvent — real appointments, created and edited here
#   • Task.due_date — overlaid read-only, edited in the وظایف tab
#   • Reminder.remind_at — overlaid read-only, edited in the یادآورها tab
# Overlay rows carry kind="task"/"reminder" so the UI knows not to offer edit.

async def _attach_event_serials(db: AsyncSession, events) -> None:
    """Stamp each event with its property's کد ملک.

    Appointments store property_id, but every screen an agent reads shows the
    serial, so the two must not be confused with each other.
    """
    ids = {e.property_id for e in events if e.property_id}
    if not ids:
        return
    rows = (await db.execute(
        select(Property.id, Property.serial_no).where(Property.id.in_(ids)))).all()
    by_id = {r.id: r.serial_no for r in rows}
    for e in events:
        e._property_serial = by_id.get(e.property_id)


async def _resolve_property_serial(db: AsyncSession, data: dict) -> None:
    """Turn a کد ملک typed into the form into the property_id we store.

    Without this the serial would be written straight into property_id and
    silently attach the appointment to a different property.
    """
    if "property_serial" not in data:
        return
    raw = data.get("property_serial")
    if raw in (None, "", 0):
        data["property_id"] = None
        return
    try:
        serial = int(raw)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="کد ملک نامعتبر است")
    prop_id = (await db.execute(
        select(Property.id).where(Property.serial_no == serial))).scalar_one_or_none()
    if prop_id is None:
        raise HTTPException(status_code=400, detail=f"ملکی با کد {serial} یافت نشد")
    data["property_id"] = prop_id


def _task_as_event(t: Task) -> dict:
    prio_color = {"urgent": "#f87171", "high": "#fb923c",
                  "medium": "#60a5fa", "low": "#94a3b8"}
    return {
        "id": t.id, "kind": "task", "title": t.title,
        "event_type": "task", "type_label": "وظیفه",
        "color": prio_color.get(t.priority, "#60a5fa"),
        "start_at": t.due_date.isoformat() if t.due_date else None,
        "end_at": None, "all_day": True, "location": None,
        "status": "done" if t.status == "done" else "scheduled",
        "assigned_to": t.assigned_to, "description": t.description,
    }


def _reminder_as_event(r: Reminder) -> dict:
    return {
        "id": r.id, "kind": "reminder", "title": r.title,
        "event_type": "reminder", "type_label": "یادآور", "color": "#e879f9",
        "start_at": r.remind_at.isoformat() if r.remind_at else None,
        "end_at": None, "all_day": False, "location": None,
        "status": "done" if r.is_sent else "scheduled",
        "assigned_to": None, "description": None,
    }


async def _calendar_rows(db: AsyncSession, start: datetime, end: datetime,
                         include_overlay: bool = True,
                         event_type: Optional[str] = None,
                         assigned_to: Optional[str] = None) -> List[dict]:
    """Every dated row that falls inside [start, end)."""
    q = select(CalendarEvent).where(
        CalendarEvent.start_at >= start, CalendarEvent.start_at < end)
    if event_type:
        q = q.where(CalendarEvent.event_type == event_type)
    if assigned_to:
        q = q.where(CalendarEvent.assigned_to == assigned_to)
    events = (await db.execute(q)).scalars().all()
    await _attach_event_serials(db, events)
    rows = [e.to_dict() for e in events]

    # A type filter is about appointment types, so it hides the overlays too
    if include_overlay and not event_type:
        tasks = (await db.execute(select(Task).where(
            Task.due_date >= start, Task.due_date < end))).scalars().all()
        rows += [_task_as_event(t) for t in tasks
                 if not assigned_to or t.assigned_to == assigned_to]
        if not assigned_to:
            rems = (await db.execute(select(Reminder).where(
                Reminder.remind_at >= start, Reminder.remind_at < end))).scalars().all()
            rows += [_reminder_as_event(r) for r in rems]

    rows.sort(key=lambda r: r["start_at"] or "")
    return rows


@router.get("/calendar")
async def list_calendar(
    date_from: str,
    date_to: str,
    event_type: Optional[str] = None,
    assigned_to: Optional[str] = None,
    include_overlay: bool = True,
    db: AsyncSession = Depends(get_db),
):
    """Events (plus tasks/reminders) in a date window — what a grid page needs."""
    try:
        start, end = _parse_datetime(date_from), _parse_datetime(date_to)
    except Exception:
        raise HTTPException(status_code=400, detail="بازهٔ تاریخ نامعتبر است")
    if end <= start:
        raise HTTPException(status_code=400, detail="تاریخ پایان باید بعد از شروع باشد")
    items = await _calendar_rows(db, start, end, include_overlay, event_type, assigned_to)
    return {"items": items, "total": len(items)}


@router.get("/calendar/upcoming")
async def upcoming_events(
    days: int = Query(7, ge=1, le=90),
    limit: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    """Next appointments from now — the dashboard strip and the «قرارهای پیشِ رو» box."""
    now = datetime.now()
    rows = await _calendar_rows(db, now, now + timedelta(days=days))
    rows = [r for r in rows if r.get("status") != "canceled"]
    return {"items": rows[:limit], "total": len(rows)}


@router.get("/calendar/export/excel")
async def export_calendar_excel(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Excel counterpart of the calendar (appointments only, not overlays)."""
    q = select(CalendarEvent).order_by(CalendarEvent.start_at.asc())
    for value, column, op in ((date_from, CalendarEvent.start_at, "ge"),
                              (date_to, CalendarEvent.start_at, "lt")):
        if value:
            try:
                dt = _parse_datetime(value)
                q = q.where(column >= dt if op == "ge" else column < dt)
            except Exception:
                pass
    items = (await db.execute(q.limit(5000))).scalars().all()

    status_fa = {"scheduled": "برنامه‌ریزی‌شده", "done": "انجام شد", "canceled": "لغو شد"}
    await _attach_event_serials(db, items)
    headers = ["#", "عنوان", "نوع", "کد ملک", "تاریخ", "ساعت", "محل بازدید",
               "مالک", "شماره مالک", "مشتری", "شماره مشتری",
               "کارشناس فروش", "شماره کارشناس",
               "یادآوری پیامکی", "وضعیت", "نتیجه", "توضیحات"]
    rows = [[
        e.id, e.title, e.type_label, getattr(e, "_property_serial", None), fa_date(e.start_at),
        "" if e.all_day else e.start_at.strftime("%H:%M"), e.location,
        e.owner_name or e.attendee_name, e.owner_phone or e.attendee_phone,
        e.customer_name, e.customer_phone,
        e.assigned_to, e.agent_phone,
        ("ارسال شد" if e.sms_sent else "فعال") if e.sms_reminder else "",
        status_fa.get(e.status, e.status), e.outcome, e.description,
    ] for e in items]
    return xlsx_response("calendar.xlsx", "تقویم", headers, rows)


def _apply_event_fields(event: CalendarEvent, data: dict) -> None:
    """Copy the writable fields off a request body onto an event."""
    for field in ("title", "description", "location", "outcome",
                  "owner_name", "owner_phone", "customer_name", "customer_phone",
                  "assigned_to", "agent_phone",
                  "attendee_name", "attendee_phone"):   # legacy, still accepted
        if field in data:
            setattr(event, field, data[field])
    if data.get("event_type") in CalendarEvent.EVENT_TYPES:
        event.event_type = data["event_type"]
    if data.get("status") in ("scheduled", "done", "canceled"):
        event.status = data["status"]
    if "all_day" in data:
        event.all_day = bool(data["all_day"])
    if "remind_before" in data:
        try:
            event.remind_before = int(data["remind_before"] or 0)
        except (TypeError, ValueError):
            pass
    if "sms_reminder" in data:
        event.sms_reminder = bool(data["sms_reminder"])
    for field in ("property_id", "lead_id", "customer_id", "contact_id", "deal_id"):
        if field in data:
            setattr(event, field, data[field] or None)
    if data.get("start_at"):
        try:
            new_start = _parse_datetime(data["start_at"])
        except Exception:
            raise HTTPException(status_code=400, detail="تاریخ شروع نامعتبر است")
        if event.start_at and new_start != event.start_at:
            # rescheduled — the attendee is owed a reminder for the new time
            event.sms_sent = False
        event.start_at = new_start
    if "end_at" in data:                      # explicit null clears the end time
        if data["end_at"]:
            try:
                event.end_at = _parse_datetime(data["end_at"])
            except Exception:
                raise HTTPException(status_code=400, detail="تاریخ پایان نامعتبر است")
        else:
            event.end_at = None


@router.post("/calendar")
async def create_event(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    if not (data.get("title") or "").strip():
        raise HTTPException(status_code=400, detail="عنوان قرار الزامی است")
    if not data.get("start_at"):
        raise HTTPException(status_code=400, detail="تاریخ شروع الزامی است")

    await _resolve_property_serial(db, data)
    event = CalendarEvent(title=data["title"].strip(), start_at=datetime.now())
    _apply_event_fields(event, data)
    if event.end_at and event.end_at < event.start_at:
        raise HTTPException(status_code=400, detail="پایان قرار قبل از شروع آن است")
    event.created_by = getattr(current_user, "full_name", None) or getattr(current_user, "username", None)

    # A visit without an address is a visit nobody can attend — borrow the
    # property's own address when the caller did not supply one.
    if not event.location and event.property_id:
        prop = (await db.execute(
            select(Property).where(Property.id == event.property_id))).scalar_one_or_none()
        if prop:
            event.location = prop.address or " ".join(
                filter(None, [prop.city_name, prop.district, prop.neighborhood])) or None

    db.add(event)
    await db.commit()
    await db.refresh(event)

    if event.lead_id:
        await _log_activity(db, "lead", event.lead_id, "event",
                            f"{event.type_label} ثبت شد: {event.title}", event.created_by)
    if event.customer_id:
        await _log_activity(db, "customer", event.customer_id, "event",
                            f"{event.type_label} ثبت شد: {event.title}", event.created_by)
    await _attach_event_serials(db, [event])
    return event.to_dict()


@router.get("/calendar/{event_id}")
async def get_event(event_id: int, db: AsyncSession = Depends(get_db)):
    event = (await db.execute(
        select(CalendarEvent).where(CalendarEvent.id == event_id))).scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="قرار یافت نشد")
    await _attach_event_serials(db, [event])
    return event.to_dict()


@router.patch("/calendar/{event_id}")
async def update_event(
    event_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    event = (await db.execute(
        select(CalendarEvent).where(CalendarEvent.id == event_id))).scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="قرار یافت نشد")

    old_status = event.status
    await _resolve_property_serial(db, data)
    _apply_event_fields(event, data)
    if event.end_at and event.end_at < event.start_at:
        raise HTTPException(status_code=400, detail="پایان قرار قبل از شروع آن است")
    await db.commit()
    await db.refresh(event)

    if event.status != old_status and event.lead_id:
        status_fa = {"scheduled": "برنامه‌ریزی‌شده", "done": "انجام شد", "canceled": "لغو شد"}
        actor = getattr(current_user, "full_name", None) or getattr(current_user, "username", None)
        await _log_activity(db, "lead", event.lead_id, "event",
                            f"{event.type_label} «{event.title}» → {status_fa.get(event.status)}", actor)
    await _attach_event_serials(db, [event])
    return event.to_dict()


@router.delete("/calendar/{event_id}")
async def delete_event(event_id: int, db: AsyncSession = Depends(get_db)):
    event = (await db.execute(
        select(CalendarEvent).where(CalendarEvent.id == event_id))).scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="قرار یافت نشد")
    await db.delete(event)
    await db.commit()
    return {"success": True}


@router.post("/calendar/{event_id}/sms")
async def send_event_sms(
    event_id: int,
    data: Optional[dict] = None,
    db: AsyncSession = Depends(get_db),
    current_user: Optional[User] = Depends(get_current_user_optional),
):
    """Text the attendee about this appointment now — confirmations, changes.

    Separate from the scheduled reminder: sending by hand does not consume
    the automatic one, so a confirmation today still gets a reminder tomorrow.
    """
    event = (await db.execute(
        select(CalendarEvent).where(CalendarEvent.id == event_id))).scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="قرار یافت نشد")

    override = ((data or {}).get("to") or "").strip()
    targets = [("طرف قرار", None, override)] if override else event.sms_targets()
    if not targets:
        raise HTTPException(status_code=400, detail="هیچ شماره‌ای برای این قرار ثبت نشده است")

    custom = ((data or {}).get("message") or "").strip()
    sent, failed = [], []
    for role, _name, phone in targets:
        message = custom or event.sms_text(role)
        res = await send_sms(phone, message)
        db.add(SmsLog(
            to_number=phone, message=message,
            status="sent" if res.get("success") else "failed",
            provider=res.get("provider", "kavenegar"),
            response=str(res.get("response", ""))[:2000],
            contact_id=event.contact_id,
        ))
        (sent if res.get("success") else failed).append(
            {"role": role, "phone": phone, "error": str(res.get("response", ""))[:200]})
    await db.commit()

    if not sent:
        detail = failed[0].get("error") if failed else ""
        raise HTTPException(status_code=502, detail=f"ارسال پیامک ناموفق بود: {detail}")
    return {"success": True, "sent": sent, "failed": failed}


@router.get("/match/property/{property_id}/customers")
async def match_customers_for_property(
    property_id: int,
    limit: int = Query(12, ge=1, le=50),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """«متقاضیان هم‌خوان» — who was looking for a file like this one.

    The inverse of /match/customer/{id}: a new file arrives and the question
    is who to ring, not what to show.
    """
    prop = (await db.execute(
        select(Property).where(Property.id == property_id))).scalar_one_or_none()
    if not prop:
        raise HTTPException(status_code=404, detail="ملک یافت نشد")
    items = await customers_for_property(db, prop, limit=limit)
    return {"items": items, "total": len(items),
            "source": {"id": prop.id, "title": prop.title, "serial_no": prop.serial_no}}
