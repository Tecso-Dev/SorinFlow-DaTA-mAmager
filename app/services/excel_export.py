"""
Shared Excel (xlsx) export helper — one styled writer for every list in the
panel, so each JSON export can have a matching «خروجی اکسل» button.
"""
import io
from typing import Iterable, List, Sequence

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_HEADER_BG = "1a1a2e"


def build_xlsx(title: str, headers: Sequence[str], rows: Iterable[Sequence]) -> io.BytesIO:
    """Render a styled, RTL, auto-width sheet and return it as a BytesIO."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31] or "Sheet1"
    ws.sheet_view.rightToLeft = True

    fill = PatternFill("solid", fgColor=_HEADER_BG)
    for col, head in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=head)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths: List[int] = [len(str(h)) + 4 for h in headers]
    for row in rows:
        ws.append(list(row))
        for i, val in enumerate(row):
            if i < len(widths):
                widths[i] = max(widths[i], min(len(str(val if val is not None else "")) + 3, 60))

    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def xlsx_response(filename: str, title: str, headers: Sequence[str],
                  rows: Iterable[Sequence]) -> StreamingResponse:
    """Convenience: build the sheet and wrap it in a download response."""
    buf = build_xlsx(title, headers, rows)
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def fa_date(value) -> str:
    """ISO-ish date for a datetime column (empty string when missing)."""
    return value.strftime("%Y-%m-%d %H:%M") if value else ""
